"""
Gera Excel de auditoria da ingestão (DEBUG_INGESTAO.xlsx).
"""

from __future__ import annotations

from dataclasses import asdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from motor_ingestao import MotorIngestao


def _escrever_aba(ws, titulo: str, registros: list[dict]) -> None:
    ws.title = titulo
    if not registros:
        ws.append(["Sem registros"])
        return

    headers = list(registros[0].keys())
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for registro in registros:
        ws.append([registro.get(h, "") for h in headers])

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def gerar_debug_excel(motor: MotorIngestao, caminho_saida: str) -> str:
    wb = Workbook()

    abas = [
        ("PDF Extraido", [asdict(p) for p in motor.pedidos_pdf]),
        ("XLSB Retidos", [asdict(r) for r in motor.retencoes_xlsb]),
        ("Email Eventos", [asdict(e) for e in motor.eventos_email]),
        ("Consolidado", [asdict(c) for c in motor.consolidados]),
        (
            "Resumo",
            [{"chave": k, "valor": v} for k, v in motor.resumo().items()],
        ),
        (
            "Erros Avisos",
            [
                {"tipo": "AVISO", "mensagem": a}
                for a in motor.avisos
            ]
            + [{"tipo": "ERRO", "mensagem": e} for e in motor.erros],
        ),
    ]

    primeira = True
    for nome, dados in abas:
        if primeira:
            ws = wb.active
            primeira = False
        else:
            ws = wb.create_sheet()
        _escrever_aba(ws, nome, dados)

    caminho = Path(caminho_saida)
    caminho.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho)
    return str(caminho)
