"""
Gerador do Romaneio de Carga — layout legado Excel por veículo.

Colunas: CLIENTE | PESO KG | PEDIDO | BAIRRO
Ordem dos itens: sequencia_lifo decrescente (maior seq. = carrega primeiro / LIFO).
"""

from __future__ import annotations

from datetime import date
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

COLUNAS = ("CLIENTE", "PESO KG", "PEDIDO", "BAIRRO")

VERDE_SISTEMA = "16A34A"
VERDE_CLARO = "DCFCE7"
BORDA_FINA = Side(style="thin", color="CCCCCC")


def _ordenar_itens_lifo(itens: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Maior sequencia_lifo primeiro (ordem de carregamento no baú)."""
    return sorted(
        itens,
        key=lambda item: int(str(item.get("sequencia_lifo") or "0") or 0),
        reverse=True,
    )


def _titulo_veiculo(rota: dict[str, Any]) -> str:
    nome = str(rota.get("veiculo_nome") or rota.get("veiculo_id") or "VEÍCULO").strip()
    return nome.upper()


def _escrever_bloco_veiculo(
    ws,
    linha_inicio: int,
    titulo: str,
    itens: list[dict[str, Any]],
) -> int:
    """Escreve bloco do veículo; retorna próxima linha livre."""
    fill_titulo = PatternFill("solid", fgColor=VERDE_SISTEMA)
    fill_header = PatternFill("solid", fgColor=VERDE_CLARO)
    font_titulo = Font(color="FFFFFF", bold=True, size=12)
    font_header = Font(bold=True, size=11)
    font_total = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    borda = Border(left=BORDA_FINA, right=BORDA_FINA, top=BORDA_FINA, bottom=BORDA_FINA)

    row = linha_inicio

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell_titulo = ws.cell(row=row, column=1, value=titulo)
    cell_titulo.fill = fill_titulo
    cell_titulo.font = font_titulo
    cell_titulo.alignment = align_center
    row += 1

    for col_idx, header in enumerate(COLUNAS, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center
        cell.border = borda
    row += 1

    peso_total = 0.0
    for item in itens:
        peso = float(item.get("peso_kg") or 0)
        peso_total += peso
        valores = (
            str(item.get("cliente") or ""),
            peso,
            str(item.get("numero_pedido") or ""),
            str(item.get("bairro_destino") or ""),
        )
        for col_idx, valor in enumerate(valores, start=1):
            cell = ws.cell(row=row, column=col_idx, value=valor)
            cell.alignment = align_left if col_idx in (1, 4) else align_center
            cell.border = borda
            if col_idx == 2:
                cell.number_format = "#,##0.00"
        row += 1

    cell_rotulo = ws.cell(row=row, column=1, value="TOTAL DE PESO")
    cell_rotulo.font = font_total
    cell_rotulo.alignment = Alignment(horizontal="right", vertical="center")
    cell_rotulo.border = borda
    cell_rotulo.fill = fill_header

    cell_peso = ws.cell(row=row, column=2, value=peso_total)
    cell_peso.font = font_total
    cell_peso.number_format = "#,##0.00"
    cell_peso.alignment = align_center
    cell_peso.border = borda
    cell_peso.fill = fill_header

    for col in range(3, 5):
        cell = ws.cell(row=row, column=col, value="")
        cell.border = borda
        cell.fill = fill_header

    return row + 2


def gerar_romaneio_xlsx(roteirizacao: dict[str, Any]) -> bytes:
    """Gera bytes do .xlsx a partir do payload roteirizacao.json."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Romaneio"

    rotas = roteirizacao.get("rotas") or []
    itens_por_veiculo: dict[str, list] = roteirizacao.get("itens_por_veiculo") or {}

    ws.merge_cells("A1:D1")
    titulo_doc = ws.cell(row=1, column=1, value=f"ROMANEIO DE CARGA — {date.today().strftime('%d/%m/%Y')}")
    titulo_doc.font = Font(bold=True, size=14)
    titulo_doc.alignment = Alignment(horizontal="center", vertical="center")

    linha = 3
    blocos = 0

    for rota in rotas:
        veiculo_id = str(rota.get("veiculo_id") or "")
        itens_raw = itens_por_veiculo.get(veiculo_id) or []
        if not itens_raw:
            continue

        itens = _ordenar_itens_lifo(list(itens_raw))
        linha = _escrever_bloco_veiculo(ws, linha, _titulo_veiculo(rota), itens)
        blocos += 1

    if blocos == 0:
        ws.cell(row=3, column=1, value="Nenhum pedido alocado em veículos.")

    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 28

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def nome_arquivo_romaneio(session_id: str = "") -> str:
    sufixo = session_id[:8] if session_id else "export"
    return f"romaneio_{sufixo}_{date.today().isoformat()}.xlsx"
